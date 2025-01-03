import unittest
import os
from src.analyzer import ExcelAnalyzer
from src.models import CellError, ErrorSeverity
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook

class TestExcelAnalyzer(unittest.TestCase):
    def setUp(self):
        self.analyzer = ExcelAnalyzer()
        self.test_files_dir = os.path.join(os.path.dirname(__file__), 'test_files')
        os.makedirs(self.test_files_dir, exist_ok=True)

    def test_analyze_file_not_found(self):
        """Test handling of non-existent file"""
        with self.assertRaises(FileNotFoundError):
            self.analyzer.analyze_file("nonexistent.xlsx")

    def test_analyze_invalid_file(self):
        """Test handling of invalid file"""
        invalid_file = os.path.join(self.test_files_dir, 'invalid.xlsx')
        with open(invalid_file, 'w') as f:
            f.write('Not a valid Excel file')
        
        with self.assertRaises(InvalidFileException):
            self.analyzer.analyze_file(invalid_file)

    def test_analyze_empty_file(self):
        """Test analyzing an empty Excel file"""
        empty_file = os.path.join(self.test_files_dir, 'empty.xlsx')
        wb = Workbook()
        wb.save(empty_file)
        
        errors = self.analyzer.analyze_file(empty_file)
        self.assertEqual(len(errors), 0)

    def test_analyze_corrupted_zip(self):
        """Test handling of corrupted ZIP structure"""
        corrupted_file = os.path.join(self.test_files_dir, 'corrupted.xlsx')
        # Create a file that starts with ZIP signature but is corrupted
        with open(corrupted_file, 'wb') as f:
            f.write(b'PK\x03\x04corrupted')
        
        with self.assertRaises(InvalidFileException):
            self.analyzer.analyze_file(corrupted_file)

    def test_error_severity_and_suggestions(self):
        """Test error severity levels and fix suggestions"""
        test_file = os.path.join(self.test_files_dir, 'test.xlsx')
        wb = Workbook()
        ws = wb.active
        
        # Add a very long string
        ws['A1'] = 'x' * 40000
        
        # Add invalid sheet name (using valid name but will be detected by our analyzer)
        wb.create_sheet('Sheet1_very_long_name_that_exceeds_limit_31_chars')
        
        wb.save(test_file)
        
        errors = self.analyzer.analyze_file(test_file)
        
        # Check error severity
        self.assertTrue(any(e.severity == ErrorSeverity.ERROR for e in errors))
        
        # Check fix suggestions
        self.assertTrue(any(e.fix_suggestion is not None for e in errors))

    def test_error_fix_suggestions(self):
        """Test error fix suggestions"""
        test_file = os.path.join(self.test_files_dir, 'suggestions.xlsx')
        wb = Workbook()
        ws = wb.active
        
        # Add a very long string
        long_string = 'x' * 40000
        cell = ws['A1']
        # Set string value using openpyxl's internal method
        cell._value = long_string
        cell.data_type = 's'  # Use 's' for string type
        
        # Add cell with zero-width space
        cell = ws['B1']
        text_with_zero_width = 'test\u200Btext'
        cell._value = text_with_zero_width
        cell.data_type = 's'
        
        # Force save with string values
        wb.save(test_file)
        
        # Run analysis with verbose mode to see what's happening
        errors = self.analyzer.analyze_file(test_file, verbose=True)
        
        # Debug print
        if not errors:
            self.fail("No errors found at all")
        for error in errors:
            print(f"\nFound error: {error.error_type} in {error.sheet_name} at {error.column}{error.row}")
            print(f"Details: {error.details}")
        
        # Find long string error
        long_string_errors = [e for e in errors if e.error_type == "Long string"]
        self.assertTrue(len(long_string_errors) > 0, "No long string error found")
        long_string_error = long_string_errors[0]
        self.assertIsNotNone(long_string_error.fix_suggestion)
        self.assertIn("Split", long_string_error.fix_suggestion)
        
        # Find special character error
        special_char_errors = [e for e in errors if e.error_type == "Special character"]
        self.assertTrue(len(special_char_errors) > 0, "No special character error found")
        special_char_error = special_char_errors[0]
        self.assertIsNotNone(special_char_error.fix_suggestion)
        self.assertIn("Remove", special_char_error.fix_suggestion)

    def tearDown(self):
        # Clean up test files
        if os.path.exists(self.test_files_dir):
            for file in os.listdir(self.test_files_dir):
                os.remove(os.path.join(self.test_files_dir, file))
            os.rmdir(self.test_files_dir) 